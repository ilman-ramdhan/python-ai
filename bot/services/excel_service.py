import logging
from io import BytesIO
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

logger = logging.getLogger("bot.services.excel")

class ExcelService:
    """Service for handling Excel file generation"""

    @staticmethod
    def generate(data: Dict[str, Any]) -> tuple[BytesIO, str]:
        """
        Creates an Excel file from structured dictionary data.
        
        Args:
            data: Dictionary containing 'sheets' list. 
                  Example structure:
                  {
                      "filename": "output.xlsx",
                      "sheets": [
                          {
                              "name": "Sheet1",
                              "headers": ["Col1", "Col2"],
                              "rows": [["Val1", "Val2"], ["Val3", "Val4"]]
                          }
                      ]
                  }
                  
        Returns:
            tuple: (BytesIO stream of the excel file, filename string)
        """
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            default_sheet = workbook.active
            if default_sheet:
                workbook.remove(default_sheet)
                
            sheets_data = data.get("sheets", [])
            if not sheets_data:
                # Fallback if no sheets provided
                ws = workbook.create_sheet("Data")
                ws.append(["No data provided"])
                
            for sheet_info in sheets_data:
                sheet_name = sheet_info.get("name", "Sheet")
                headers = sheet_info.get("headers", [])
                rows = sheet_info.get("rows", [])
                
                # Create worksheet
                ws: Worksheet = workbook.create_sheet(title=sheet_name)
                
                # Add headers
                if headers:
                    ws.append(headers)
                    # Style headers (bold)
                    for cell in ws[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                
                # Add data rows
                for row in rows:
                    ws.append(row)
                    
                # Auto-adjust column widths
                for i, col in enumerate(ws.columns, 1):
                    max_length = 0
                    column = get_column_letter(i)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap width at 50

            # Save to memory
            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)
            
            filename = data.get("filename", "generated_data.xlsx")
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
                
            return excel_file, filename

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}", exc_info=True)
            raise e
