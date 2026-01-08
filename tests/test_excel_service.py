import pytest
from io import BytesIO
from bot.services.excel_service import ExcelService
import openpyxl

def test_create_excel_basic():
    """Test basic excel creation with valid data"""
    data = {
        "filename": "test.xlsx",
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["Name", "Age"],
                "rows": [["Alice", 30], ["Bob", 25]]
            }
        ]
    }
    
    excel_io, filename = ExcelService.generate(data)
    
    assert filename == "test.xlsx"
    assert isinstance(excel_io, BytesIO)
    
    # Verify content
    wb = openpyxl.load_workbook(excel_io)
    sheet = wb["Sheet1"]
    assert sheet["A1"].value == "Name"
    assert sheet["B1"].value == "Age"
    assert sheet["A2"].value == "Alice"
    assert sheet["B2"].value == 30

def test_create_excel_no_extension():
    """Test filename auto-correction"""
    data = {
        "filename": "report",
        "sheets": []
    }
    _, filename = ExcelService.generate(data)
    assert filename == "report.xlsx"

def test_create_excel_empty():
    """Test empty data handling"""
    data = {}
    excel_io, _ = ExcelService.generate(data)
    
    wb = openpyxl.load_workbook(excel_io)
    assert "Data" in wb.sheetnames
    assert wb["Data"]["A1"].value == "No data provided"
